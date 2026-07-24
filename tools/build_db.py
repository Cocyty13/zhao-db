#!/usr/bin/env python3
"""Build a fast, lossless SQLite snapshot from the Zhao Excel workbook."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sqlite3
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


COMMON_INDEX_COLUMNS = {
    "ID",
    "Scene_ID",
    "Event_ID",
    "Character",
    "Name",
    "Canon_Status",
    "Status",
    "Current_State",
    "Last_Scene",
}


def quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def sql_name(value: str, fallback: str) -> str:
    name = re.sub(r"[^0-9A-Za-z_]+", "_", value.strip()).strip("_")
    if not name:
        name = fallback
    if name[0].isdigit():
        name = f"c_{name}"
    return name


def cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def used_rows(sheet: Any) -> list[list[Any]]:
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    while rows and not any(value is not None and str(value) != "" for value in rows[-1]):
        rows.pop()
    if not rows:
        return []
    width = max(
        (
            index + 1
            for row in rows
            for index, value in enumerate(row)
            if value is not None and str(value) != ""
        ),
        default=0,
    )
    return [row[:width] + [None] * max(0, width - len(row)) for row in rows]


def unique_columns(headers: Iterable[Any]) -> list[tuple[str, str]]:
    seen: dict[str, int] = {}
    result: list[tuple[str, str]] = []
    for position, raw in enumerate(headers, start=1):
        original = cell_text(raw) or f"Column_{position}"
        base = sql_name(original, f"column_{position}")
        seen[base] = seen.get(base, 0) + 1
        sql_column = base if seen[base] == 1 else f"{base}__{seen[base]}"
        result.append((original, sql_column))
    return result


def create_support_tables(connection: sqlite3.Connection) -> bool:
    connection.executescript(
        """
        CREATE TABLE _meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE _sheets (
            position INTEGER PRIMARY KEY,
            sheet_name TEXT UNIQUE NOT NULL,
            table_name TEXT UNIQUE NOT NULL,
            source_rows INTEGER NOT NULL,
            data_rows INTEGER NOT NULL,
            column_count INTEGER NOT NULL
        );
        CREATE TABLE _columns (
            sheet_name TEXT NOT NULL,
            position INTEGER NOT NULL,
            original_header TEXT NOT NULL,
            sqlite_column TEXT NOT NULL,
            PRIMARY KEY (sheet_name, position)
        );
        """
    )
    try:
        connection.execute(
            "CREATE VIRTUAL TABLE search_index USING fts5("
            "sheet_name UNINDEXED, row_number UNINDEXED, record_id UNINDEXED, content)"
        )
        return True
    except sqlite3.OperationalError:
        connection.execute(
            "CREATE TABLE search_index ("
            "sheet_name TEXT, row_number INTEGER, record_id TEXT, content TEXT)"
        )
        connection.execute(
            "CREATE INDEX idx_search_sheet_row ON search_index(sheet_name, row_number)"
        )
        return False


def build(source: Path, destination: Path) -> dict[str, Any]:
    workbook = load_workbook(source, data_only=False, read_only=False)
    destination.unlink(missing_ok=True)
    Path(f"{destination}-wal").unlink(missing_ok=True)
    Path(f"{destination}-shm").unlink(missing_ok=True)
    connection = sqlite3.connect(destination)
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("PRAGMA foreign_keys=ON")
    fts_enabled = create_support_tables(connection)

    summary: list[dict[str, Any]] = []
    with connection:
        connection.executemany(
            "INSERT INTO _meta(key, value) VALUES (?, ?)",
            [
                ("format_version", "1"),
                ("source_file", source.name),
                ("generated_at_utc", dt.datetime.now(dt.timezone.utc).isoformat()),
                ("sheet_count", str(len(workbook.sheetnames))),
                ("fts5_enabled", "true" if fts_enabled else "false"),
            ],
        )

        for sheet_position, sheet_name in enumerate(workbook.sheetnames, start=1):
            rows = used_rows(workbook[sheet_name])
            if not rows:
                headers: list[tuple[str, str]] = []
                data: list[list[Any]] = []
            else:
                headers = unique_columns(rows[0])
                data = rows[1:]

            table_name = sql_name(sheet_name, f"sheet_{sheet_position}")
            column_sql = ", ".join(
                f"{quote(sql_column)} TEXT" for _, sql_column in headers
            )
            suffix = f", {column_sql}" if column_sql else ""
            connection.execute(
                f"CREATE TABLE {quote(table_name)} "
                f"(_row_number INTEGER PRIMARY KEY{suffix})"
            )

            connection.execute(
                "INSERT INTO _sheets("
                "position, sheet_name, table_name, source_rows, data_rows, column_count"
                ") VALUES (?, ?, ?, ?, ?, ?)",
                (
                    sheet_position,
                    sheet_name,
                    table_name,
                    len(rows),
                    len(data),
                    len(headers),
                ),
            )
            connection.executemany(
                "INSERT INTO _columns("
                "sheet_name, position, original_header, sqlite_column"
                ") VALUES (?, ?, ?, ?)",
                [
                    (sheet_name, index, original, sql_column)
                    for index, (original, sql_column) in enumerate(headers, start=1)
                ],
            )

            if headers:
                insert_columns = ", ".join(
                    ["_row_number"] + [quote(sql_column) for _, sql_column in headers]
                )
                placeholders = ", ".join("?" for _ in range(len(headers) + 1))
                prepared_rows = [
                    [excel_row] + [cell_text(value) for value in row]
                    for excel_row, row in enumerate(data, start=2)
                ]
                connection.executemany(
                    f"INSERT INTO {quote(table_name)} ({insert_columns}) "
                    f"VALUES ({placeholders})",
                    prepared_rows,
                )

                header_map = {
                    original: sql_column for original, sql_column in headers
                }
                for original, sql_column in headers:
                    if original in COMMON_INDEX_COLUMNS:
                        index_name = sql_name(
                            f"idx_{table_name}_{sql_column}", "idx_lookup"
                        )
                        connection.execute(
                            f"CREATE INDEX {quote(index_name)} "
                            f"ON {quote(table_name)} ({quote(sql_column)})"
                        )

                record_column = next(
                    (
                        header_map[name]
                        for name in ("ID", "Scene_ID", "Event_ID", "Name", "Character")
                        if name in header_map
                    ),
                    None,
                )
                search_rows = []
                for excel_row, row in enumerate(data, start=2):
                    values = [cell_text(value) for value in row]
                    content = " | ".join(value for value in values if value)
                    record_id = (
                        values[[c for _, c in headers].index(record_column)]
                        if record_column
                        else None
                    )
                    search_rows.append(
                        (sheet_name, excel_row, record_id, content)
                    )
                connection.executemany(
                    "INSERT INTO search_index("
                    "sheet_name, row_number, record_id, content"
                    ") VALUES (?, ?, ?, ?)",
                    search_rows,
                )

            summary.append(
                {
                    "sheet_name": sheet_name,
                    "table_name": table_name,
                    "source_rows": len(rows),
                    "data_rows": len(data),
                    "columns": len(headers),
                }
            )

        connection.execute("PRAGMA optimize")
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]

    connection.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    connection.execute("PRAGMA journal_mode=DELETE")
    connection.close()
    return {
        "source": str(source),
        "destination": str(destination),
        "fts5_enabled": fts_enabled,
        "integrity_check": integrity,
        "sheets": summary,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    parser.add_argument("--summary", type=Path)
    args = parser.parse_args()

    result = build(args.source, args.destination)
    payload = json.dumps(result, ensure_ascii=False, indent=2)
    if args.summary:
        args.summary.write_text(payload + "\n", encoding="utf-8")
    print(payload)


if __name__ == "__main__":
    main()
